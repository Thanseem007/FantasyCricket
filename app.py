import os
import json
from flask import Flask, render_template, request, redirect, url_for, flash,send_file,session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user 
from google.cloud import storage
import openpyxl
from google.cloud import storage
import io
from flask import jsonify
from filelock import FileLock
import logging
from cloud_logger import CloudLogger
from cloud_logger import setup_logger
from io import BytesIO
from datetime import datetime, timedelta
import csv
from google.cloud.exceptions import PreconditionFailed
import time

BUCKET_NAME = "fantasy_cricket_asia_1" 
DO_SERVER_DOWN = "doServerdown.txt"
USER_TABLE = "userdata.xlsx"
LOG_FILE_NAME = "log_file.txt"
LEADERBOARD_FILE = "leaderboard.csv"
PLAYERLEADERBOARD_FILE = "playerleaderboard.csv"

app = Flask(__name__)
app.secret_key = "secretkey"  # Replace with a secure key


# Configure Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

team_dir = "user_teams"  # Directory to store user teams
is_Cloud = True
if is_Cloud :
   storage_client = storage.Client()
   #Cloud Logger
   cloud_logger = CloudLogger(BUCKET_NAME, LOG_FILE_NAME)
   cloud_logger.setLevel(logging.INFO)
   cloud_logger.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

   # Add the CloudLogger handler to the Flask app's logger
   app.logger = setup_logger()
   app.logger.addHandler(cloud_logger)
   app.logger.setLevel(logging.INFO)
   app.logger.info("Hello world endpoint hit.")



# Ensure the directory exists
os.makedirs(team_dir, exist_ok=True)

# Load player data from JSON
def load_players():
    try:
        with open("players.json", "r") as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

players = load_players()

class User(UserMixin):
    def __init__(self, username):
        self.id = username

@login_manager.user_loader
def load_user(username):
    return User(username)

@app.route("/Index")
@login_required
def index():
    username = current_user.id
    team = load_team(username)
    team_player_ids = [player["id"] for player in team]
    budget = 100 - sum(player["price"] for player in team)
    return render_template("index.html", players=players, budget=budget, team=team,team_player_ids= team_player_ids,originalTeam=originalTeam)

@app.route("/" , methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = str(request.form["username"])
        apartment_number = str(request.form["apartment_number"])
          
        # Fetch credentials from the Excel file
        credentials = get_credentials_from_excel()
        # Check if the username and password exist
        user_exists = any(str(user) == username and str(pwd) == apartment_number for user, pwd in credentials)
        if not user_exists:
                flash("Incorrect username or PIN.", category="error")
                return redirect(url_for("login"))

        user = User(username)
        login_user(user)
        
        # Save apartment number
        session['apartment_number'] = apartment_number

        # Load team if it exists
        team = load_team(username)
        if not team:  # If no team, initialize an empty team
            save_team(username, [])
        if IsDownTimeReached(username) :
            return redirect(url_for("server_down"))
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/server_down")
def server_down():
    return render_template("server_down.html")

@app.route("/select/<int:player_id>")
@login_required
def select_player(player_id):
    username = current_user.id
    team = load_team(username)
    player = next((p for p in players if p["id"] == player_id), None)

    if player and player not in team:
        budget = 100 - sum(p["price"] for p in team)
        if budget >= player["price"] and len(team) < 11:
            team.append(player)
            save_team(username, team)  # Save the updated team
    return redirect(url_for("index"))

@app.route("/remove/<int:player_id>")
@login_required
def remove_player(player_id):
    username = current_user.id
    team = load_team(username)
    player = next((p for p in team if p["id"] == player_id), None)

    if player:
        team.remove(player)
        save_team(username, team)  # Save the updated team
    return redirect(url_for("index"))

@app.route("/submit", methods=["POST"])
@login_required
def submit_team():
    username = current_user.id
    app.logger.info(f"Submitting team for user {username}")
    # Lock file path to prevent race conditions during file update
    lock_filename = "/tmp/user_players.lock"  # The lock file path on your system
    lock = FileLock(lock_filename,timeout=10)
    
    max_retries = 3  # Maximum retry attempts
    backoff_factor = 2  # Exponential backoff factor

    # Retry logic for the team submission process
    attempt = 0
    while attempt < max_retries:
        try:
            if IsDownTimeReached(username) :
                return redirect(url_for("server_down"))
            username = current_user.id
            apartment_number = session.get('apartment_number', 'Unknown')
            team = load_team(username)
    
            captain_id = request.form.get('captain_id')
    # file_url = blob.public_url  # Make the file publicly accessible
            with lock:
                app.logger.info(f"Acquired lock to update Excel file for user {username}")
                file_url = update_excel_file(username, apartment_number,team,captain_id) 
            app.logger.info(f"Team submitted successfully for user {username}.") 
            return render_template("submit.html", team=team)

        except PreconditionFailed:
            
            attempt += 1
            app.logger.error(f"PreconditionFailed while updating Excel file for user {username}. Attempt {attempt}/{max_retries}")

            if attempt < max_retries:
                sleep_time = backoff_factor ** attempt
                app.logger.info(f"Retrying attempt {attempt}/{max_retries} after {sleep_time} seconds...")
                time.sleep(sleep_time)
            else:
                app.logger.error(f"Max retries reached for user {username}. Could not update file.")
                return render_template("error.html", error_message="An unexpected error occurred. Please try again later.")
        
        except Exception as e:
            app.logger.error(f"Error occurred while submitting team for user {username}: {str(e)}")
            return render_template("error.html", error_message="An unexpected error occurred. Please try again later.")
    

@app.route('/update_captain/<int:player_id>', methods=['POST'])
def update_captain(player_id):
    try:
        # Get player ID from the request
        #player_id = request.json.get('player_id')

        if not player_id:
            return jsonify({'error': 'Player ID is required'}), 400


        # Read the file from Google Cloud Storage
        username = current_user.id
        team = load_team(username)

        # Update IsCaptain field
        captain_updated = False
        for player in team:
            if player["id"] == player_id:
                player["IsCaptain"] = 1
                captain_updated = True
            else:
                player["IsCaptain"] = 0

        if not captain_updated:
            return jsonify({'error': 'Player not found'}), 404

        # Save the updated JSON
        save_team(username, team)

        return jsonify({'message': 'Captain updated successfully'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/toggle_flag', methods=['GET'])
def toggle_flag():
    # Retrieve the file from the bucket
    bucket = get_storage_client().bucket(BUCKET_NAME)
    blob = bucket.blob(DO_SERVER_DOWN)
    
    # Download the current content of the file
    current_value =  blob.download_as_text().strip()

    # Update the content (for example, replacing the current value)
    new_value = '1' if current_value == '0' else '0'
    
    # Upload the updated content back to the bucket
    blob.upload_from_string(new_value)
    
    return jsonify({'message': f'File updated successfully with value {new_value}'}), 200

@app.route('/send_credentials', methods=['GET'])
def Send_Credentials():
    try:
        # Load the Excel file with openpyxl
        workbook = openpyxl.load_workbook("userdatamail.xlsx")
        sheet = workbook.active

        # Iterate over rows and send emails
        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            username, password, recipient_email = row
            msg = GetMessage(username, password, recipient_email)
            send(msg)
            #results.append(result)

        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route('/download_team', methods=['POST'])
def download_team():
    # Get the team data from the request
    data = request.get_json()
    team = data.get('team', [])

    if not team:
        return jsonify({"error": "No team data provided"}), 400

    # Create a StringIO object to store the text data
    output = BytesIO()
    utc_now = datetime.utcnow()
    # Manually adjust the time to IST (UTC + 5:30)
    ist_now = utc_now + timedelta(hours=5, minutes=30)
    timestamp = ist_now.strftime("%Y-%m-%d %H:%M:%S")
    # Write timestamp to the output (in binary format, using encode)
    output.write(f"Team generated on: {timestamp}\n\n".encode('utf-8')) 
    # Write player names to the StringIO object
    for player in team:
        if player['IsCaptain'] == 1 :
          output.write((player['name'] + '(C) \n').encode('utf-8'))
        else :
          output.write((player['name'] + '\n').encode('utf-8'))
    # Set the cursor to the beginning of the file
    output.seek(0)
    

    # Send the file as a response
    return send_file(output, as_attachment=True, download_name="team_members.txt", mimetype="text/plain")


@app.route("/leaderboard")
def leaderboard():
    leaderboard_data = read_leaderboard()
    return render_template("leaderboard.html", leaderboard=leaderboard_data)


@app.route("/playerleaderboard")
def playerleaderboard():
    leaderboard_data = read_playerleaderboard()
    return render_template("playerleaderboard.html", leaderboard=leaderboard_data)

# Sample data for matches
matches = [
    {
        "date": "2025-01-04",
        "time": "20:00",
        "team1": "GOA",
        "team2": "Pattaya Pangalis",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184443/prestige-premier-league-s7/gumbal-of-alaparais-goa-vs-pattaya-pankalis-pp"
    },
    {
        "date": "2025-01-04",
        "time": "22:30",
        "team1": "Pondy Pulaingos",
        "team2": "Thailand Thugs",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184450/prestige-premier-league-s7/thailand-thugs-tt-vs-pondy-pulaingos-pop"
    },
    {
        "date": "2025-01-11",
        "time": "20:00",
        "team1": "Thailand Thugs",
        "team2": "Pondy Pulaingos",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184475/prestige-premier-league-s7/thailand-thugs-tt-vs-pondy-pulaingos-pop"
    },
    {
        "date": "2025-01-11",
        "time": "22:30",
        "team1": "Vegas Veriyans",
        "team2": "GOA",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184491/prestige-premier-league-s7/gumbal-of-alaparais-goa-vs-vegas-veriyans-vv"
    },
    {
        "date": "2025-01-24",
        "time": "22:30",
        "team1": "Vegas Veriyans",
        "team2": "The Varkala King",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184502/prestige-premier-league-s7/the-varkala-kings-tvk-vs-vegas-veriyans-vv"
    },
    {
        "date": "2025-01-25",
        "time": "22:30",
        "team1": "Pattaya Pangalis",
        "team2": "Pondy Pulaingos",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184525/prestige-premier-league-s7/pattaya-pankalis-pp-vs-pondy-pulaingos-pop"
    },
    {
        "date": "2025-01-31",
        "time": "22:30",
        "team1": "GOA",
        "team2": "Thailand Thugs",
        "venue": "Tavya,Chennai",
         "link":"https://cricheroes.com/scorecard/14184530/prestige-premier-league-s7/thailand-thugs-tt-vs-gumbal-of-alaparais-goa"
    },
    {
        "date": "2025-02-01",
        "time": "22:30",
        "team1": "The Varkala King",
        "team2": "Pattaya Pangalis",
        "venue": "Tavya,Chennai",
         "link":"https://cricheroes.com/scorecard/14184560/prestige-premier-league-s7/the-varkala-kings-tvk-vs-pattaya-pankalis-pp"
    },
    {
        "date": "2025-02-08",
        "time": "20:00",
        "team1": "Vegas Veriyans",
        "team2": "Pondy Pulaingos",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184566/prestige-premier-league-s7/pondy-pulaingos-pop-vs-vegas-veriyans-vv"
    },
    {
        "date": "2025-02-08",
        "time": "22:30",
        "team1": "The Varkala King",
        "team2": "GOA",
        "venue": "Tavya,Chennai",
        "link":"https://cricheroes.com/scorecard/14184568/prestige-premier-league-s7/gumbal-of-alaparais-goa-vs-the-varkala-kings-tvk"
    },
    {
        "date": "2025-02-15",
        "time": "22:30",
        "team1": "Thailand Thugs",
        "team2": "Pattaya Pangalis",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-02-22",
        "time": "20:00",
        "team1": "The Varkala King",
        "team2": "Vegas Veriyans",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-02-22",
        "time": "22:30",
        "team1": "Pondy Pulaingos",
        "team2": "GOA",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-01",
        "time": "20:00",
        "team1": "Pondy Pulaingos",
        "team2": "Vegas Veriyans",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-01",
        "time": "22:30",
        "team1": "Thailand Thugs",
        "team2": "The Varkala King",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-08",
        "time": "20:00",
        "team1": "GOA",
        "team2": "The Varkala King",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-08",
        "time": "22:30",
        "team1": "Pondy Pulaingos",
        "team2": "Pattaya Pangalis",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-15",
        "time": "20:00",
        "team1": "Thailand Thugs",
        "team2": "Pattaya Pangalis",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-15",
        "time": "22:30",
        "team1": "Vegas Veriyans",
        "team2": "GOA",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-22",
        "time": "20:00",
        "team1": "Pondy Pulaingos",
        "team2": "The Varkala King",
        "venue": "Tavya,Chennai",
    },
    {
        "date": "2025-03-22",
        "time": "22:30",
        "team1": "Vegas Veriyans",
        "team2": "Thailand Thugs",
        "venue": "Tavya,Chennai",
    }
]


originalTeam = ["PP","POP","VV","TVK","TT","GOA"]

@app.route('/upcoming_matches')
def upcomingmatches():
    today = datetime.today()
    end_of_week = today + timedelta(days=7)
    for match in matches:
     match_datetime = datetime.strptime(f"{match['date']} {match['time']}", "%Y-%m-%d %H:%M")
     match['status'] = "Completed" if match_datetime < datetime.now() else "Upcoming"
     match_date = datetime.strptime(match["date"], "%Y-%m-%d")
     match["day"] = match_date.strftime("%A") 
     match["datetime"] = match_datetime
    
    # Split matches into two categories
    completed_matches = [ match for match in matches if today >= match["datetime"] ]
    
    matches_this_week = [
        match for match in matches if today <= match["datetime"] <= end_of_week
    ]

    other_matches = [
        match for match in matches if match["datetime"]  > end_of_week
    ]

    return render_template("matchdetails.html", matches_this_week=matches_this_week, other_matches=other_matches,completed_matches=completed_matches)



# Helper Functions
def save_team(username, team):
    apartment_number = session.get('apartment_number', 'Unknown')
    if is_Cloud :
      filename = f"{username}_{apartment_number}.json"
      client = get_storage_client()
      bucket = client.bucket(BUCKET_NAME)
      blob = bucket.blob(filename)
      blob.upload_from_string(json.dumps(team))
    else :
      filepath = os.path.join(team_dir, f"{username}_{apartment_number}.json")
      with open(filepath, "w") as file:
          json.dump(team, file)

def load_team(username):
    apartment_number = session.get('apartment_number', 'Unknown')
    filename = f"{username}_{apartment_number}.json"
    if(is_Cloud) :
       client = get_storage_client()
       bucket = client.bucket(BUCKET_NAME)
       blob = bucket.blob(filename)
       if blob.exists():
          return json.loads(blob.download_as_text())
    else :
        filepath = os.path.join(team_dir, f"{username}_{apartment_number}.json")
        if os.path.exists(filepath):
           with open(filepath, "r") as file:
            return json.load(file)
    return []

def get_storage_client():
    return storage_client

def get_storage_client_new():
    return storage.Client()

def update_excel_file(username,apartment_number ,team,captain_id):
    # Define the Excel file name in the GCS bucket
    excel_filename = "user_players.xlsx"
    
    # Initialize the GCS client and bucket
    client = get_storage_client()
    bucket = client.bucket(BUCKET_NAME)  # Replace with your actual GCS bucket name
    blob = bucket.blob(excel_filename)

    # Check if the file exists
    if blob.exists():
        # Download the existing file from GCS
        blob.reload() 
        excel_data = blob.download_as_bytes()
        generation_number = blob.generation
        workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
        sheet = workbook.active
    else:
        # If the file does not exist, create a new Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Player 1", "Player 2", "Player 3", "Player 4", "Player 5", "Player 6", "Player 7", "Player 8", "Player 9", "Player 10", "Player 11", "Captain","PIN"])  # Add header row
    # Prepare the row to be added (username + selected players)
    row = [username]  # Start with the username in the first column

    # Add player names to the row
    captain= ""
    for player in team:
        if str(player["id"]) == str(captain_id):
            row.append(player["name"])
            captain=player["name"]
        else:
            row.append(player["name"])

    # Append the row to the sheet
    #row.extend([""] * (12 - len(row)))  # Fill remaining columns if fewer than 11 players
    row.append(captain)
    row.append(apartment_number)  # Add the apartment number as the last column

    # Check if an entry for this username and apartment number already exists
    existing_row = None
    app.logger.info(f"User {username}.Apartment {apartment_number}") 
    for r_idx, existing_row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if str(existing_row_data[0]) == str(username) and str(existing_row_data[-1]) == str(apartment_number):
            existing_row = r_idx
            break

    if existing_row:
        # Update the existing row
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=existing_row, column=col_idx, value=value)
    else:
        # Append a new row if no existing entry is found
        sheet.append(row)

    # Save the workbook to a BytesIO object
    updated_excel_data = io.BytesIO()
    workbook.save(updated_excel_data)
    updated_excel_data.seek(0)  # Rewind the stream to the beginning
    
    try:
     
     if generation_number is None:
                    # Upload without if_generation_match for new file
        blob.upload_from_file(updated_excel_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
     else :
        current_generation = blob.generation
        app.logger.info(f"Generation1 - {current_generation}") 
     # Upload the updated Excel file back to GCS
        blob.upload_from_file(updated_excel_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", if_generation_match=current_generation)
    except PreconditionFailed:
         raise PreconditionFailed("The file has been updated by another process. Please retry.")

    
    # Make the file publicly accessible (optional)
    blob.make_public()

    # Return the public URL of the Excel file (or you can return a signed URL if needed)
    return blob.public_url

def IsDownTimeReached(username):
    try:
        if str(username) ==str("Test"):
          return False
        if(is_Cloud) :
        # Initialize a client
            client = get_storage_client()

        # Get the bucket
            bucket = client.bucket(BUCKET_NAME)

        # Get the blob (file)
            blob = bucket.blob(DO_SERVER_DOWN)

        # Read the content of the file
            flag_value = blob.download_as_text().strip()
        else :
            filepath = os.path.join(team_dir, DO_SERVER_DOWN)
            with open(filepath, 'r') as file:
              flag_value = file.read().strip()
        # Check if the flag is 0
        print(type(flag_value))
        if flag_value =="1" :
         return True
        return False
    except Exception as e:
        print(f"An error occurred: {e}")


def get_credentials_from_excel():
    """Fetch username and password from Excel in Google Cloud Storage bucket."""
    # Initialize a Google Cloud Storage client
    if is_Cloud :
     client = get_storage_client()
     bucket = client.get_bucket(BUCKET_NAME)
     blob = bucket.blob(USER_TABLE)

    # Download the file content as bytes
     excel_data = blob.download_as_bytes()

    # Read the Excel file into a pandas DataFrame
     workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
     sheet = workbook.active  # Assuming credentials are in the first sheet

     credentials = []
     for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        username, password = row[:2]  # Assuming 'username' and 'password' are in the first two columns
        credentials.append((username, password))
    else :
     wb = openpyxl.load_workbook(USER_TABLE)
     sheet = wb.active
     credentials = []
     for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        username, password = row[:2]  # Assuming 'username' and 'password' are in the first two columns
        credentials.append((username, password))
    return credentials

def read_leaderboard():
    """Reads leaderboard data from a CSV file and sorts it by score."""
    leaderboard = []
    try:
        if is_Cloud :
            storage_client = storage.Client()
            bucket = storage_client.bucket(BUCKET_NAME)
            blob = bucket.blob(LEADERBOARD_FILE)

            # Download the CSV data as a string
            csv_data = blob.download_as_text()
            # Parse the CSV data
            reader = csv.DictReader(io.StringIO(csv_data))
            reader.fieldnames = [header.strip().lower() for header in reader.fieldnames]
            for row in reader:
                leaderboard.append({
                "name": row["name"],
                "score": int(row["score"])  # Convert score to integer for sorting
            })
        else :
            with open(LEADERBOARD_FILE, mode="r") as file:
                reader = csv.DictReader(file)
                print(reader.fieldnames)
                reader.fieldnames = [header.strip().lower() for header in reader.fieldnames]
                print(reader)
                for row in reader:
                    leaderboard.append({
                "name": row["name"],
                "score": int(row["score"])  # Convert score to integer for sorting
                })
        # Sort by score in descending order
        sorted_leaderboard = sorted(leaderboard, key=lambda x: x["score"], reverse=True)
        # Assign ranks dynamically
        for i, player in enumerate(sorted_leaderboard, start=1):
            player["rank"] = i
    except FileNotFoundError:
        print("File not found. Please ensure the leaderboard file is present.")
        return []
    except KeyError as e:
        print(f"Missing column in CSV: {e}")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []
    return sorted_leaderboard


def read_playerleaderboard():
    """Reads leaderboard data from a CSV file and sorts it by score."""
    leaderboard = []
    try:
        if is_Cloud :
            storage_client = storage.Client()
            bucket = storage_client.bucket(BUCKET_NAME)
            blob = bucket.blob(PLAYERLEADERBOARD_FILE)

            # Download the CSV data as a string
            csv_data = blob.download_as_text()
            # Parse the CSV data
            reader = csv.DictReader(io.StringIO(csv_data))
            reader.fieldnames = [header.strip().lower() for header in reader.fieldnames]
            for row in reader:
                leaderboard.append({
                "name": row["name"],
                "score": int(row["score"])  # Convert score to integer for sorting
            })
        else :
            with open(PLAYERLEADERBOARD_FILE, mode="r") as file:
                reader = csv.DictReader(file)
                print(reader.fieldnames)
                reader.fieldnames = [header.strip().lower() for header in reader.fieldnames]
                print(reader)
                for row in reader:
                    leaderboard.append({
                "name": row["name"],
                "score": int(row["score"])  # Convert score to integer for sorting
                })
        # Sort by score in descending order
        sorted_leaderboard = sorted(leaderboard, key=lambda x: x["score"], reverse=True)
        # Assign ranks dynamically
        for i, player in enumerate(sorted_leaderboard, start=1):
            player["rank"] = i
    except FileNotFoundError:
        print("File not found. Please ensure the leaderboard file is present.")
        return []
    except KeyError as e:
        print(f"Missing column in CSV: {e}")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []
    return sorted_leaderboard



if __name__ == "__main__":
    app.run(debug=True)
